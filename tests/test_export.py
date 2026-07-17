"""The full-school export: complete, current, tenant-isolated.

The export is the school's insurance policy — "our records belong to us." The
most important property is the one nobody sees: one school's workbook must
never contain another school's children.
"""
import io

from openpyxl import load_workbook

from tests.conftest import headers


def _book(content: bytes):
    return load_workbook(io.BytesIO(content), read_only=True)


async def test_export_contains_every_sheet_and_the_actual_data(ctx):
    client, ids = ctx
    ah = await headers(client, "admin@gss-ikeja.ng", "admin123")

    # put some marks and money in so the export has content
    comp = ids["component_ids"]
    await client.post("/api/scores", headers=ah, json={
        "subject_id": ids["subject_id"], "arm_id": ids["arm_id"],
        "term_id": ids["term_id"],
        "rows": [{"student_id": ids["student_ids"][0], "scores": {comp[3]: 68}}]})
    await client.post("/api/results/compute", headers=ah,
                      json={"arm_id": ids["arm_id"], "term_id": ids["term_id"]})

    r = await client.get("/api/export/school.xlsx", headers=ah)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert "fiyox-school-export" in r.headers["content-disposition"]

    wb = _book(r.content)
    for sheet in ("Students", "Guardians", "Staff", "Results",
                  "Subject scores", "Invoices", "Payments", "Attendance"):
        assert sheet in wb.sheetnames, f"missing sheet: {sheet}"

    students = list(wb["Students"].values)
    assert students[0][:4] == ("Admission No", "First name", "Last name", "Gender")
    names = {f"{row[1]} {row[2]}" for row in students[1:]}
    assert "Chinedu Eze" in names

    results = list(wb["Results"].values)
    assert len(results) >= 2            # header + at least the computed student
    chinedu = next(row for row in results[1:] if row[3] == "Chinedu Eze")
    assert chinedu[4] == 68             # grand total made it into the file


async def test_export_is_admin_only(ctx):
    client, ids = ctx
    th = await headers(client, "teacher@gss-ikeja.ng", "teach123")
    assert (await client.get("/api/export/school.xlsx", headers=th)).status_code == 403


async def test_export_never_leaks_another_schools_children(ctx):
    client, ids = ctx

    # a second school with its own admin and student
    from app.core.security import hash_password
    from app.models.school import Role, School, User
    from app.models.student import Student
    async with ids["session_factory"]() as db:
        other = School(name="Rival College", slug="rival-college")
        db.add(other)
        await db.flush()
        db.add(User(school_id=other.id, email="admin@rival.ng",
                    hashed_password=hash_password("rival123"),
                    role=Role.SCHOOL_ADMIN, first_name="Rival", last_name="Admin"))
        db.add(Student(school_id=other.id, admission_number="RIV/26/001",
                       first_name="Secret", last_name="Child", gender="female"))
        await db.commit()

    # school A's export: no rival child anywhere in the workbook
    ah = await headers(client, "admin@gss-ikeja.ng", "admin123")
    wb = _book((await client.get("/api/export/school.xlsx", headers=ah)).content)
    all_text = " ".join(str(c) for ws in wb.worksheets
                        for row in ws.values for c in row if c)
    assert "Secret Child" not in all_text
    assert "RIV/26/001" not in all_text

    # and the rival's own export contains ONLY their child
    rh = await headers(client, "admin@rival.ng", "rival123")
    rwb = _book((await client.get("/api/export/school.xlsx", headers=rh)).content)
    rstudents = list(rwb["Students"].values)[1:]
    assert [row[0] for row in rstudents] == ["RIV/26/001"]


async def test_exporting_leaves_an_audit_trail(ctx):
    client, ids = ctx
    ah = await headers(client, "admin@gss-ikeja.ng", "admin123")
    await client.get("/api/export/school.xlsx", headers=ah)

    logs = (await client.get("/api/audit-logs", headers=ah,
            params={"table_name": "school"})).json()
    assert any(l["action"] == "export" for l in logs), \
        "a full export of every child and payment must leave a trace"
