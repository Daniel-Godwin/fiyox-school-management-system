"""Full school data export — the answer to "what happens to our records?"

A school's data belongs to the school. This produces one Excel workbook with a
sheet per record type, readable by anyone with Excel or Google Sheets, needing
no knowledge of Fiyox at all:

    Students · Guardians · Staff · Results · Subject scores · Invoices ·
    Payments · Attendance

Design decisions:
* IDs are included as columns (they make re-import or dispute resolution
  possible) but pushed to the far right so the human-readable columns lead.
* Every sheet is plain tabular data — no formulas, no macros — so the file is
  safe to open anywhere and will still be readable in twenty years.
* The export is tenant-scoped like everything else: one school, its data only.
"""
import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.academics import ClassArm, SchoolClass, Subject, Term
from app.models.attendance import Attendance
from app.models.fees import Invoice, Payment
from app.models.results import SubjectResult, TermResult
from app.models.school import User
from app.models.student import Guardian, Student

HEADER_FILL = PatternFill("solid", fgColor="1D3557")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _sheet(wb: Workbook, title: str, headers: list[str], rows: list[list]):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r in rows:
        ws.append(r)
    # sensible column widths: header length or longest value, capped
    for idx, h in enumerate(headers, start=1):
        longest = max([len(str(h))] + [len(str(r[idx - 1])) for r in rows[:200]
                                       if idx - 1 < len(r) and r[idx - 1] is not None])
        ws.column_dimensions[get_column_letter(idx)].width = min(max(longest + 2, 10), 40)
    ws.freeze_panes = "A2"
    return ws


async def build_school_export(db: AsyncSession, school_id: str) -> bytes:
    """One workbook containing everything the school would need to walk away."""
    # ---- shared lookups -------------------------------------------------
    arms = {a.id: a for a in (await db.execute(select(ClassArm).where(
        ClassArm.school_id == school_id))).scalars().all()}
    classes = {c.id: c.name for c in (await db.execute(select(SchoolClass).where(
        SchoolClass.school_id == school_id))).scalars().all()}
    terms = {t.id: t for t in (await db.execute(select(Term).where(
        Term.school_id == school_id))).scalars().all()}
    subjects = {s.id: s.name for s in (await db.execute(select(Subject).where(
        Subject.school_id == school_id))).scalars().all()}
    users = {u.id: u for u in (await db.execute(select(User).where(
        User.school_id == school_id))).scalars().all()}
    students = {s.id: s for s in (await db.execute(select(Student).where(
        Student.school_id == school_id,
        Student.deleted_at.is_(None)))).scalars().all()}

    def arm_label(aid: str | None) -> str:
        a = arms.get(aid) if aid else None
        return f"{classes.get(a.class_id, '')} {a.name}".strip() if a else ""

    def term_label(tid: str | None) -> str:
        t = terms.get(tid) if tid else None
        return f"{str(t.name).split('.')[-1].title()} term" if t else ""

    def student_name(sid: str | None) -> str:
        s = students.get(sid) if sid else None
        return f"{s.first_name} {s.last_name}" if s else ""

    wb = Workbook()
    wb.remove(wb.active)   # drop the default empty sheet

    # ---- Students -------------------------------------------------------
    _sheet(wb, "Students",
           ["Admission No", "First name", "Last name", "Gender", "Class",
            "Active", "Student ID"],
           [[s.admission_number, s.first_name, s.last_name,
             str(getattr(s.gender, "value", s.gender) or ""),
             arm_label(s.current_arm_id),
             "Yes" if s.is_active else "No", s.id]
            for s in sorted(students.values(), key=lambda x: x.admission_number)])

    # ---- Guardians -------------------------------------------------------
    links = (await db.execute(select(Guardian).where(
        Guardian.school_id == school_id))).scalars().all()
    _sheet(wb, "Guardians",
           ["Parent name", "Parent email", "Parent phone", "Phone verified",
            "Child", "Child admission no", "Relationship"],
           [[(f"{users[l.parent_user_id].first_name} {users[l.parent_user_id].last_name}"
              if l.parent_user_id in users else ""),
             users[l.parent_user_id].email if l.parent_user_id in users else "",
             users[l.parent_user_id].phone if l.parent_user_id in users else "",
             ("Yes" if l.parent_user_id in users
              and users[l.parent_user_id].phone_verified else "No"),
             student_name(l.student_id),
             students[l.student_id].admission_number if l.student_id in students else "",
             l.relationship or ""]
            for l in links])

    # ---- Staff ------------------------------------------------------------
    _sheet(wb, "Staff",
           ["Name", "Email", "Phone", "Role", "Active"],
           [[f"{u.first_name} {u.last_name}", u.email, u.phone or "",
             str(getattr(u.role, "value", u.role)), "Yes" if u.is_active else "No"]
            for u in users.values()
            if str(getattr(u.role, "value", u.role)) in
            ("school_admin", "teacher", "bursar")])

    # ---- Results (term summaries) -----------------------------------------
    trs = (await db.execute(select(TermResult).where(
        TermResult.school_id == school_id))).scalars().all()
    _sheet(wb, "Results",
           ["Term", "Class", "Admission No", "Student", "Total", "Average",
            "Position", "Class size", "Class average", "Form teacher's comment",
            "Principal's comment"],
           [[term_label(r.term_id), arm_label(r.arm_id),
             students[r.student_id].admission_number if r.student_id in students else "",
             student_name(r.student_id), r.grand_total, r.average,
             r.overall_position, r.class_size, r.class_average,
             r.form_teacher_comment or "", r.principal_comment or ""]
            for r in trs])

    # ---- Subject scores -----------------------------------------------------
    srs = (await db.execute(select(SubjectResult).where(
        SubjectResult.school_id == school_id))).scalars().all()
    _sheet(wb, "Subject scores",
           ["Term", "Admission No", "Student", "Subject", "Total", "Grade",
            "Position", "Class average"],
           [[term_label(r.term_id),
             students[r.student_id].admission_number if r.student_id in students else "",
             student_name(r.student_id), subjects.get(r.subject_id, ""),
             r.total, r.grade, r.subject_position, r.class_average]
            for r in srs])

    # ---- Invoices ------------------------------------------------------------
    invoices = (await db.execute(select(Invoice).where(
        Invoice.school_id == school_id,
        Invoice.deleted_at.is_(None)))).scalars().all()
    _sheet(wb, "Invoices",
           ["Invoice No", "Term", "Admission No", "Student", "Amount",
            "Discount", "Status", "Due date", "Invoice ID"],
           [[i.invoice_number, term_label(i.term_id),
             students[i.student_id].admission_number if i.student_id in students else "",
             student_name(i.student_id), i.amount, i.discount,
             str(getattr(i.status, "value", i.status)),
             str(i.due_date) if i.due_date else "", i.id]
            for i in invoices])

    # ---- Payments ---------------------------------------------------------------
    payments = (await db.execute(select(Payment).where(
        Payment.school_id == school_id,
        Payment.deleted_at.is_(None)))).scalars().all()
    inv_by_id = {i.id: i for i in invoices}
    _sheet(wb, "Payments",
           ["Reference", "Invoice No", "Student", "Amount", "Method",
            "Paid on", "Received by"],
           [[p.reference,
             inv_by_id[p.invoice_id].invoice_number if p.invoice_id in inv_by_id else "",
             student_name(inv_by_id[p.invoice_id].student_id)
             if p.invoice_id in inv_by_id else "",
             p.amount, str(getattr(p.method, "value", p.method)),
             str(p.paid_at) if p.paid_at else "",
             (f"{users[p.received_by].first_name} {users[p.received_by].last_name}"
              if p.received_by and p.received_by in users else "")]
            for p in payments])

    # ---- Attendance ---------------------------------------------------------------
    att = (await db.execute(select(Attendance).where(
        Attendance.school_id == school_id))).scalars().all()
    _sheet(wb, "Attendance",
           ["Date", "Admission No", "Student", "Class", "Status"],
           [[str(a.date),
             students[a.student_id].admission_number if a.student_id in students else "",
             student_name(a.student_id), arm_label(a.arm_id),
             str(getattr(a.status, "value", a.status))]
            for a in sorted(att, key=lambda x: str(x.date))])

    # ---- About sheet (provenance) --------------------------------------------------
    about = wb.create_sheet("About this export", 0)
    about["A1"] = "Fiyox School Management System — full data export"
    about["A1"].font = Font(bold=True, size=13)
    about["A3"] = "Exported at (UTC):"
    about["B3"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")
    about["A4"] = "Sheets:"
    about["B4"] = ("Students, Guardians, Staff, Results, Subject scores, "
                   "Invoices, Payments, Attendance")
    about["A6"] = ("This file is plain tabular data — no formulas or macros — and "
                   "opens in Excel, Google Sheets, or LibreOffice. Your school's "
                   "records belong to your school.")
    about.column_dimensions["A"].width = 22
    about.column_dimensions["B"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
